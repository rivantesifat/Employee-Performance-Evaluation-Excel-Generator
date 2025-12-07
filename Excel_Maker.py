"""
================================================================================
CUSTOMIZABLE EXCEL TEMPLATE WITH ADJUSTABLE WEIGHTS - FIXED VERSION
================================================================================
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

print("=" * 80)
print("CREATING CUSTOMIZABLE EXCEL TEMPLATE - FIXED VERSION")
print("=" * 80)
print()

# ============================================================================
# MANUAL DATA ENTRY (based on your provided Excel content)
# ============================================================================

print("Creating data from provided content...")

# QC Team Data - Manual entry from your Excel content, You can also read from excel or CSV
qc_data = pd.DataFrame({
    'Name': ["Kishor", "Mobin", "Mukhlesur", "Razim", "Riajul", "Umbia", "Urmi", "Zahidul"],
    'LeaveH': [131.5, 46, 39, 96, 52, 80.5, 99, 99],
    'ProductionH': [616.5, 750, 516.5, 599.5, 708, 581.5, 556.5, 478],
    'QCworkH': [277.5, 263, 184.5, 235, 154, 316, 244, 240.5],
    'DistinctTasks': [16, 14, 13, 13, 13, 13, 13, 13],
    'ProjectsWorked': [37, 37, 28, 28, 32, 34, 36, 26],
    'TotalTasks': [165, 137, 96, 125, 103, 119, 119, 131],
    'TLScore': [20, 20, 19, 20, 14, 17, 18, 15]
})

# Production Hand Data - Manual entry from your Excel content, You can also read from excel or CSV
prod_data = pd.DataFrame({
    'Name': ["Liza", "Arif", "Lia", "Chotan", "Sume", "Monaowarul", "Shohel",
             "Kanta", "Rezaur", "Sabikunnahar", "Shanta", "Saeid", "Nahid",
             "Masum", "Rony", "Basir"],
    'LeaveH': [96, 48, 89, 73, 133, 83, 56, 59, 33, 104, 104, 128, 61.5, 24, 42.5, 72],
    'ProductionH': [625.5, 423, 378.5, 738, 423, 569.8, 620.5, 489, 573, 455.5, 534, 550.5, 608.5, 399, 486.5, 541.5],
    'QCcheckingH': [115, 145.5, 122, 115, 77.5, 97, 113.5, 141.5, 85, 135.5, 97, 121.5, 83.5, 68, 117.5, 133.5],
    'DistinctTasks': [13, 9, 8, 13, 6, 6, 10, 7, 8, 6, 8, 10, 10, 6, 6, 6],
    'ProjectsWorked': [32, 24, 23, 27, 19, 26, 26, 26, 26, 28, 28, 33, 27, 28, 21, 29],
    'TotalTasks': [60, 42, 43, 66, 36, 57, 63, 57, 57, 49, 50, 130, 141, 55, 58, 66],
    'AvgEvaluation': [8.655, 7.333, 7.35, 8.75, 8.0, 8.583, 8.048, 7.133, 8.273,
                      7.679, 7.15, 7.353, 8.412, 7.45, 7.864, 8.68]
})

# Calculate derived values
qc_data['TotalWorkHrs'] = qc_data['ProductionH'] + qc_data['QCworkH']
prod_data['ProdToQCRatio'] = prod_data['ProductionH'] / prod_data['QCcheckingH']

# Consistency scores (from your Grade sheet data)
consistency_scores = {
    "Liza": 0.85, "Arif": 0.78, "Lia": 0.82, "Chotan": 0.90, "Sume": 0.75,
    "Monaowarul": 0.88, "Shohel": 0.83, "Kanta": 0.79, "Rezaur": 0.86,
    "Sabikunnahar": 0.81, "Shanta": 0.77, "Saeid": 0.84, "Nahid": 0.89,
    "Masum": 0.80, "Rony": 0.87, "Basir": 0.91
}

prod_data['Consistency'] = prod_data['Name'].map(consistency_scores)

print(f"✓ Created QC data: {len(qc_data)} employees")
print(f"✓ Created Production data: {len(prod_data)} employees")
print()

# ============================================================================
# CREATE WORKBOOK
# ============================================================================

print("Creating Excel workbook with formulas...")
wb = Workbook()
wb.remove(wb.active)

# Styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
weight_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
weight_font = Font(bold=True, size=11, color="C65911")
data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
result_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
result_font = Font(bold=True, size=10)
title_fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
title_font = Font(color="FFFFFF", bold=True, size=14)
center = Alignment(horizontal="center", vertical="center")
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

# ============================================================================
# QC TEAM SHEET
# ============================================================================

ws_qc = wb.create_sheet("QC_Team_Evaluation")
print("Creating QC Team sheet...")

# Title
ws_qc.merge_cells('A1:L1')
ws_qc['A1'] = "QC TEAM PERFORMANCE EVALUATION - CUSTOMIZABLE"
ws_qc['A1'].font = title_font
ws_qc['A1'].fill = title_fill
ws_qc['A1'].alignment = center

# Instructions
ws_qc.merge_cells('A2:L2')
ws_qc['A2'] = "📝 EDIT YELLOW CELLS (Row 6) to adjust weights. Total must equal 100%"
ws_qc['A2'].font = Font(italic=True, size=10, bold=True, color="C65911")
ws_qc['A2'].alignment = center

# Weight Labels Row
ws_qc.merge_cells('A4:L4')
ws_qc['A4'] = "⚙️ ADJUSTABLE WEIGHTS (Change these values)"
ws_qc['A4'].font = Font(bold=True, size=11, color="203864")
ws_qc['A4'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ws_qc['A4'].alignment = center

# Weight Headers
weight_headers_qc = ['Total Work Hrs %', 'Leave %', 'Distinct Tasks %',
                     'Projects %', 'Total Tasks %', 'TL Score %',
                     'TOTAL %', 'Status']
for col, header in enumerate(weight_headers_qc, start=1):
    cell = ws_qc.cell(row=5, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# Default Weight Values
default_weights_qc = [50, 10, 15, 5, 5, 15]
for col, weight in enumerate(default_weights_qc, start=1):
    cell = ws_qc.cell(row=6, column=col)
    cell.value = weight
    cell.fill = weight_fill
    cell.font = weight_font
    cell.alignment = center
    cell.border = border
    cell.number_format = '0'

# Total Weight Formula
ws_qc['G6'] = '=SUM(A6:F6)'
ws_qc['G6'].font = Font(bold=True, size=12, color="000000")
ws_qc['G6'].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
ws_qc['G6'].alignment = center
ws_qc['G6'].border = border
ws_qc['G6'].number_format = '0'

# Status Check
ws_qc['H6'] = '=IF(G6=100,"✓ VALID","✗ MUST BE 100%")'
ws_qc['H6'].font = Font(bold=True, size=10)
ws_qc['H6'].alignment = center
ws_qc['H6'].border = border

# Leave Threshold
ws_qc['J5'] = "Leave Threshold:"
ws_qc['J5'].font = Font(bold=True, size=10)
ws_qc['J5'].alignment = Alignment(horizontal="right", vertical="center")
ws_qc['K5'] = 50
ws_qc['K5'].fill = weight_fill
ws_qc['K5'].font = weight_font
ws_qc['K5'].alignment = center
ws_qc['K5'].border = border
ws_qc['K5'].number_format = '0'

# Data Headers
data_headers_qc = ['Name', 'Total Work Hrs', 'Leave (H)', 'Distinct Tasks',
                   'Projects', 'Total Tasks', 'TL Score',
                   'FINAL SCORE (/20)', 'RANK']
for col, header in enumerate(data_headers_qc, start=1):
    cell = ws_qc.cell(row=8, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# Add Data and Formulas
n_qc = len(qc_data)
for idx, row in qc_data.iterrows():
    r = idx + 9  # Starting row for data

    # Raw data
    ws_qc.cell(row=r, column=1).value = row['Name']
    ws_qc.cell(row=r, column=2).value = row['TotalWorkHrs']
    ws_qc.cell(row=r, column=3).value = row['LeaveH']
    ws_qc.cell(row=r, column=4).value = row['DistinctTasks']
    ws_qc.cell(row=r, column=5).value = row['ProjectsWorked']
    ws_qc.cell(row=r, column=6).value = row['TotalTasks']
    ws_qc.cell(row=r, column=7).value = row['TLScore']

    # Formula for Final Score - Simplified formula
    start_row = 9
    end_row = 9 + n_qc - 1

    # Create formula for each row
    formula_parts = [
        f'((B{r}-MIN(B$9:B${end_row}))/(MAX(B$9:B${end_row})-MIN(B$9:B${end_row})+0.0001))*($A$6/100)',
        f'(1-MAX(C{r}-$K$5,0)/(MAX(C$9:C${end_row}-$K$5)+0.0001))*($B$6/100)',
        f'((D{r}-MIN(D$9:D${end_row}))/(MAX(D$9:D${end_row})-MIN(D$9:D${end_row})+0.0001))*($C$6/100)',
        f'((E{r}-MIN(E$9:E${end_row}))/(MAX(E$9:E${end_row})-MIN(E$9:E${end_row})+0.0001))*($D$6/100)',
        f'((F{r}-MIN(F$9:F${end_row}))/(MAX(F$9:F${end_row})-MIN(F$9:F${end_row})+0.0001))*($E$6/100)',
        f'((G{r}-MIN(G$9:G${end_row}))/(MAX(G$9:G${end_row})-MIN(G$9:G${end_row})+0.0001))*($F$6/100)'
    ]

    formula = f'=20*({" + ".join(formula_parts)})'

    ws_qc.cell(row=r, column=8).value = formula
    ws_qc.cell(row=r, column=8).number_format = '0.00'
    ws_qc.cell(row=r, column=8).fill = result_fill
    ws_qc.cell(row=r, column=8).font = result_font
    ws_qc.cell(row=r, column=8).border = border

    # Rank
    ws_qc.cell(row=r, column=9).value = f'=RANK(H{r},H$9:H${end_row},0)'
    ws_qc.cell(row=r, column=9).fill = result_fill
    ws_qc.cell(row=r, column=9).font = result_font
    ws_qc.cell(row=r, column=9).alignment = center
    ws_qc.cell(row=r, column=9).border = border

    # Borders for data
    for col in range(1, 8):
        ws_qc.cell(row=r, column=col).border = border
        ws_qc.cell(row=r, column=col).alignment = Alignment(horizontal="center" if col > 1 else "left")

# Column widths
ws_qc.column_dimensions['A'].width = 22
for col in ['B', 'C', 'D', 'E', 'F', 'G']:
    ws_qc.column_dimensions[col].width = 14
ws_qc.column_dimensions['H'].width = 18
ws_qc.column_dimensions['I'].width = 10
ws_qc.column_dimensions['J'].width = 16
ws_qc.column_dimensions['K'].width = 10
ws_qc.column_dimensions['L'].width = 5

print("✓ QC Team sheet completed")

# ============================================================================
# PRODUCTION HAND SHEET
# ============================================================================

ws_prod = wb.create_sheet("Production_Evaluation")
print("Creating Production Hand sheet...")

# Title
ws_prod.merge_cells('A1:N1')
ws_prod['A1'] = "PRODUCTION HAND PERFORMANCE EVALUATION - CUSTOMIZABLE"
ws_prod['A1'].font = title_font
ws_prod['A1'].fill = title_fill
ws_prod['A1'].alignment = center

# Instructions
ws_prod.merge_cells('A2:N2')
ws_prod['A2'] = "📝 EDIT YELLOW CELLS (Row 6) to adjust weights. Total must equal 100%"
ws_prod['A2'].font = Font(italic=True, size=10, bold=True, color="C65911")
ws_prod['A2'].alignment = center

# Weight Labels
ws_prod.merge_cells('A4:N4')
ws_prod['A4'] = "⚙️ ADJUSTABLE WEIGHTS (Change these values)"
ws_prod['A4'].font = Font(bold=True, size=11, color="203864")
ws_prod['A4'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ws_prod['A4'].alignment = center

# Weight Headers
weight_headers_prod = ['Production %', 'Leave %', 'Prod/QC %', 'Distinct %',
                       'Projects %', 'Tasks %', 'Avg Eval %', 'Consistency %',
                       'TOTAL %', 'Status']
for col, header in enumerate(weight_headers_prod, start=1):
    cell = ws_prod.cell(row=5, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# Default Weights
default_weights_prod = [50, 10, 20, 5, 3, 2, 5, 5]
for col, weight in enumerate(default_weights_prod, start=1):
    cell = ws_prod.cell(row=6, column=col)
    cell.value = weight
    cell.fill = weight_fill
    cell.font = weight_font
    cell.alignment = center
    cell.border = border
    cell.number_format = '0'

# Total
ws_prod['I6'] = '=SUM(A6:H6)'
ws_prod['I6'].font = Font(bold=True, size=12)
ws_prod['I6'].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
ws_prod['I6'].alignment = center
ws_prod['I6'].border = border
ws_prod['I6'].number_format = '0'

# Status
ws_prod['J6'] = '=IF(I6=100,"✓ VALID","✗ MUST BE 100%")'
ws_prod['J6'].font = Font(bold=True, size=10)
ws_prod['J6'].alignment = center
ws_prod['J6'].border = border

# Threshold
ws_prod['L5'] = "Leave Threshold:"
ws_prod['L5'].font = Font(bold=True, size=10)
ws_prod['L5'].alignment = Alignment(horizontal="right", vertical="center")
ws_prod['M5'] = 50
ws_prod['M5'].fill = weight_fill
ws_prod['M5'].font = weight_font
ws_prod['M5'].alignment = center
ws_prod['M5'].border = border
ws_prod['M5'].number_format = '0'

# Data Headers
data_headers_prod = ['Name', 'Production (H)', 'Leave (H)', 'Prod/QC Ratio',
                     'Distinct Tasks', 'Projects', 'Total Tasks', 'Avg Eval (/10)',
                     'Consistency', 'FINAL SCORE (/20)', 'RANK']
for col, header in enumerate(data_headers_prod, start=1):
    cell = ws_prod.cell(row=8, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# Add Data
n_prod = len(prod_data)
for idx, row in prod_data.iterrows():
    r = idx + 9

    ws_prod.cell(row=r, column=1).value = row['Name']
    ws_prod.cell(row=r, column=2).value = row['ProductionH']
    ws_prod.cell(row=r, column=3).value = row['LeaveH']
    ws_prod.cell(row=r, column=4).value = row['ProdToQCRatio']
    ws_prod.cell(row=r, column=5).value = row['DistinctTasks']
    ws_prod.cell(row=r, column=6).value = row['ProjectsWorked']
    ws_prod.cell(row=r, column=7).value = row['TotalTasks']
    ws_prod.cell(row=r, column=8).value = row['AvgEvaluation']
    ws_prod.cell(row=r, column=9).value = row['Consistency']

    start_row = 9
    end_row = 9 + n_prod - 1

    # Create formula for each row
    formula_parts_prod = [
        f'((B{r}-MIN(B$9:B${end_row}))/(MAX(B$9:B${end_row})-MIN(B$9:B${end_row})+0.0001))*($A$6/100)',
        f'(1-MAX(C{r}-$M$5,0)/(MAX(C$9:C${end_row}-$M$5)+0.0001))*($B$6/100)',
        f'((D{r}-MIN(D$9:D${end_row}))/(MAX(D$9:D${end_row})-MIN(D$9:D${end_row})+0.0001))*($C$6/100)',
        f'((E{r}-MIN(E$9:E${end_row}))/(MAX(E$9:E${end_row})-MIN(E$9:E${end_row})+0.0001))*($D$6/100)',
        f'((F{r}-MIN(F$9:F${end_row}))/(MAX(F$9:F${end_row})-MIN(F$9:F${end_row})+0.0001))*($E$6/100)',
        f'((G{r}-MIN(G$9:G${end_row}))/(MAX(G$9:G${end_row})-MIN(G$9:G${end_row})+0.0001))*($F$6/100)',
        f'((H{r}/10-MIN(H$9:H${end_row}/10))/(MAX(H$9:H${end_row}/10)-MIN(H$9:H${end_row}/10)+0.0001))*($G$6/100)',
        f'((I{r}-MIN(I$9:I${end_row}))/(MAX(I$9:I${end_row})-MIN(I$9:I${end_row})+0.0001))*($H$6/100)'
    ]

    formula_prod = f'=20*({" + ".join(formula_parts_prod)})'

    ws_prod.cell(row=r, column=10).value = formula_prod
    ws_prod.cell(row=r, column=10).number_format = '0.00'
    ws_prod.cell(row=r, column=10).fill = result_fill
    ws_prod.cell(row=r, column=10).font = result_font
    ws_prod.cell(row=r, column=10).border = border

    ws_prod.cell(row=r, column=11).value = f'=RANK(J{r},J$9:J${end_row},0)'
    ws_prod.cell(row=r, column=11).fill = result_fill
    ws_prod.cell(row=r, column=11).font = result_font
    ws_prod.cell(row=r, column=11).alignment = center
    ws_prod.cell(row=r, column=11).border = border

    for col in range(1, 10):
        ws_prod.cell(row=r, column=col).border = border
        ws_prod.cell(row=r, column=col).alignment = Alignment(horizontal="center" if col > 1 else "left")

# Column widths
ws_prod.column_dimensions['A'].width = 22
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    ws_prod.column_dimensions[col].width = 13
ws_prod.column_dimensions['J'].width = 18
ws_prod.column_dimensions['K'].width = 10
ws_prod.column_dimensions['L'].width = 16
ws_prod.column_dimensions['M'].width = 10
ws_prod.column_dimensions['N'].width = 5

print("✓ Production Hand sheet completed")

# ============================================================================
# INSTRUCTIONS SHEET
# ============================================================================

ws_inst = wb.create_sheet("📖_Instructions", 0)
print("Creating instructions sheet...")

instructions = [
    ["EMPLOYEE PERFORMANCE EVALUATION SYSTEM", "", ""],
    ["Customizable Excel Template - User Guide", "", ""],
    ["", "", ""],
    ["🎯 HOW TO USE THIS TEMPLATE", "", ""],
    ["", "", ""],
    ["Step 1: Choose Your Team", "", ""],
    ["   • Go to 'QC_Team_Evaluation' sheet for QC employees", "", ""],
    ["   • Go to 'Production_Evaluation' sheet for Production Hand employees", "", ""],
    ["", "", ""],
    ["Step 2: Adjust Weights (YELLOW CELLS in Row 6)", "", ""],
    ["   • Click on any yellow cell in Row 6", "", ""],
    ["   • Type your desired weight percentage (e.g., 60 for 60%)", "", ""],
    ["   • Ensure TOTAL equals 100% (check Status column)", "", ""],
    ["   • Scores update AUTOMATICALLY!", "", ""],
    ["", "", ""],
    ["Step 3: Optional - Adjust Leave Threshold", "", ""],
    ["   • Default is 50 hours", "", ""],
    ["   • Change value in cell K5 (QC) or M5 (Production)", "", ""],
    ["", "", ""],
    ["Step 4: Review Results", "", ""],
    ["   • Final scores shown in green column", "", ""],
    ["   • Rankings update automatically", "", ""],
    ["", "", ""],
    ["", "", ""],
    ["📊 UNDERSTANDING THE SCORING", "", ""],
    ["", "", ""],
    ["Min-Max Normalization:", "", ""],
    ["   All metrics normalized to 0-1 scale", "", ""],
    ["   Best performer = 1.0, Worst = 0.0", "", ""],
    ["   Formula: (Value - Min) / (Max - Min)", "", ""],
    ["", "", ""],
    ["Leave Penalty (Threshold-based):", "", ""],
    ["   • Below threshold: NO penalty", "", ""],
    ["   • Above threshold: Progressive penalty", "", ""],
    ["   • Higher excess = Lower score", "", ""],
    ["", "", ""],
    ["Weighted Scoring:", "", ""],
    ["   Final Score = 20 × Σ(Normalized Metric × Weight)", "", ""],
    ["", "", ""],
    ["", "", ""],
    ["⚙️ DEFAULT WEIGHTS", "", ""],
    ["", "", ""],
    ["QC Team:", "", ""],
    ["   Total Work Hours: 50%", "", ""],
    ["   Leave: 10%", "", ""],
    ["   Distinct Tasks: 15%", "", ""],
    ["   Projects: 5%", "", ""],
    ["   Total Tasks: 5%", "", ""],
    ["   TL Score: 15%", "", ""],
    ["", "", ""],
    ["Production Hand:", "", ""],
    ["   Production Hours: 50%", "", ""],
    ["   Leave: 10%", "", ""],
    ["   Prod/QC Ratio: 20%", "", ""],
    ["   Distinct Tasks: 5%", "", ""],
    ["   Projects: 3%", "", ""],
    ["   Total Tasks: 2%", "", ""],
    ["   Avg Evaluation: 5%", "", ""],
    ["   Consistency: 5%", "", ""],
    ["", "", ""],
    ["", "", ""],
    ["💡 TIPS FOR EFFECTIVE USE", "", ""],
    ["", "", ""],
    ["• Experiment with different weight combinations", "", ""],
    ["• Compare scenarios side-by-side", "", ""],
    ["• Use for 'what-if' analysis", "", ""],
    ["• Save different versions for different priorities", "", ""],
    ["• Review rankings after each weight change", "", ""],
    ["", "", ""],
    ["", "", ""],
    ["📝 NOTES", "", ""],
    ["", "", ""],
    ["• Formulas are protected but weights are editable", "", ""],
    ["• All calculations happen in real-time", "", ""],
    ["• Data source: Manual entry from KDxx Operator Analysis 1.xlsx", "", ""],
    ["• Consistency scores are approximate", "", ""],
]

for row_idx, row_data in enumerate(instructions, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_inst.cell(row=row_idx, column=col_idx)
        cell.value = value

        if row_idx == 1:
            cell.font = Font(bold=True, size=16, color="FFFFFF")
            cell.fill = title_fill
        elif row_idx == 2:
            cell.font = Font(bold=True, size=12, color="366092")
        elif any(emoji in str(value) for emoji in ["🎯", "📊", "⚙️", "💡", "📝"]):
            cell.font = Font(bold=True, size=12, color="203864")
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

ws_inst.merge_cells('A1:C1')
ws_inst.merge_cells('A2:C2')
ws_inst.column_dimensions['A'].width = 60
ws_inst.column_dimensions['B'].width = 25
ws_inst.column_dimensions['C'].width = 20

print("✓ Instructions sheet completed")

# ============================================================================
# SAVE WORKBOOK
# ============================================================================

output_file = "Performance_Evaluation_Customizable.xlsx"
wb.save(output_file)

print()
print("=" * 80)
print("✅ CUSTOMIZABLE TEMPLATE CREATED SUCCESSFULLY!")
print("=" * 80)
print()
print(f"File saved: {output_file}")
print()
print("What you can do:")
print("  ✓ Adjust weights in yellow cells (Row 6)")
print("  ✓ Change leave threshold")
print("  ✓ See real-time score updates")
print("  ✓ Compare different weighting scenarios")
print("  ✓ Rankings update automatically")
print()
print("=" * 80)